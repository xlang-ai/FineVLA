#!/usr/bin/env python3
"""Statistics per robot_type across LeRobot v2.0 / v2.1 datasets.

Reads all meta/info.json files, groups by robot_type, and sums
total_episodes and total duration (total_frames / fps).
Outputs a summary table as xlsx + printed to stdout.

Usage:
    python calc_robot_type_stats.py \
        --roots /path/to/Lerobot_v20 /path/to/Lerobot_v21
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

IGNORE_DATASETS = {"IPEC-COMMUNITY", "RoboCOIN_annotations_backup"}


def scan_all_info_jsons(roots: list[str]) -> list[tuple[str, str]]:
    """Return list of (info_json_path, dataset_name) tuples.

    Uses fixed-depth glob patterns instead of os.walk for speed on
    network file-systems (CPFS / NFS).
    Patterns matched (relative to each root):
        */meta/info.json        — dataset with meta at depth 1
        */*/meta/info.json      — dataset with sub-datasets at depth 2
        */*/*/meta/info.json    — depth 3 (rare but possible)
    """
    all_items: list[tuple[str, str]] = []
    seen: set[str] = set()

    for root in roots:
        root_path = Path(root)
        if not root_path.exists():
            print(f"  Warning: {root} does not exist, skipping")
            continue
        for depth_pattern in ("*/meta/info.json",
                              "*/*/meta/info.json",
                              "*/*/*/meta/info.json"):
            for fp in sorted(root_path.glob(depth_pattern)):
                fp_str = str(fp)
                if fp_str in seen:
                    continue
                dataset_name = str(fp.parent.parent.relative_to(root_path))
                top_level = dataset_name.split(os.sep)[0]
                if top_level in IGNORE_DATASETS:
                    continue
                seen.add(fp_str)
                all_items.append((fp_str, dataset_name))
    return all_items


def main():
    parser = argparse.ArgumentParser(
        description="Compute per-robot_type episode count and total hours",
    )
    parser.add_argument(
        "--roots", nargs="+", required=True,
        help="Root directories (e.g. Lerobot_v20 Lerobot_v21)",
    )
    parser.add_argument(
        "--output", type=str, default=None,
        help="Output xlsx file path (default: auto-generated with timestamp)",
    )
    args = parser.parse_args()

    if args.output is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            f"robot_type_stats_{ts}.xlsx",
        )

    wall_start = time.time()

    print("[1/3] Scanning for info.json files (glob) ...")
    info_items = scan_all_info_jsons(args.roots)
    print(f"  Found {len(info_items)} info.json files ({time.time() - wall_start:.1f}s)\n")

    if not info_items:
        print("No info.json files found. Exiting.")
        sys.exit(1)

    print("[2/3] Reading and aggregating by robot_type ...")

    def _read_info(item: tuple[str, str]) -> dict:
        fp, dataset_name = item
        try:
            with open(fp, "r", encoding="utf-8") as f:
                info = json.load(f)
            return {"path": fp, "dataset_name": dataset_name, "info": info}
        except Exception as e:
            return {"path": fp, "dataset_name": dataset_name, "error": str(e)}

    total = len(info_items)
    results: list[dict] = []
    with ThreadPoolExecutor(max_workers=32) as pool:
        futures = {pool.submit(_read_info, item): item for item in info_items}
        for i, fut in enumerate(as_completed(futures), 1):
            results.append(fut.result())
            if i % 50 == 0 or i == total:
                print(f"  read {i}/{total}", flush=True)

    stats: dict[str, dict] = defaultdict(lambda: {
        "total_episodes": 0,
        "total_frames": 0,
        "total_duration_seconds": 0.0,
        "dataset_count": 0,
        "dataset_names": [],
        "sources": [],
    })
    errors = []

    for res in results:
        if "error" in res:
            errors.append({"path": res["path"], "error": res["error"]})
            continue

        info = res["info"]
        dataset_name = res["dataset_name"]
        fp = res["path"]

        robot_type = info.get("robot_type") or "unknown"
        total_episodes = info.get("total_episodes", 0)
        total_frames = info.get("total_frames")
        fps = info.get("fps")

        if total_frames is None or fps is None:
            errors.append({"path": fp, "error": "missing total_frames or fps"})
            continue
        total_frames = int(total_frames)
        fps = float(fps)
        if fps <= 0:
            errors.append({"path": fp, "error": f"invalid fps={fps}"})
            continue

        duration_s = total_frames / fps
        entry = stats[robot_type]
        entry["total_episodes"] += int(total_episodes)
        entry["total_frames"] += total_frames
        entry["total_duration_seconds"] += duration_s
        entry["dataset_count"] += 1
        if dataset_name not in entry["dataset_names"]:
            entry["dataset_names"].append(dataset_name)
        entry["sources"].append(fp)

    rows = []
    for robot_type, s in sorted(stats.items(), key=lambda x: x[1]["total_duration_seconds"], reverse=True):
        rows.append({
            "robot_type": robot_type,
            "episode_number": s["total_episodes"],
            "total_hours": round(s["total_duration_seconds"] / 3600, 2),
            "total_frames": s["total_frames"],
            "dataset_count": s["dataset_count"],
            "dataset_names": sorted(s["dataset_names"]),
            "total_duration_seconds": s["total_duration_seconds"],
        })

    grand_episodes = sum(r["episode_number"] for r in rows)
    grand_hours = sum(r["total_hours"] for r in rows)
    grand_frames = sum(r["total_frames"] for r in rows)

    elapsed = time.time() - wall_start

    print(f"\n{'=' * 120}")
    print("  Robot Type Statistics (from info.json: total_frames / fps)")
    print(f"{'=' * 120}")
    print(f"  {'robot_type':<30s} {'episode_number':>15s} {'total_hours':>15s}  {'dataset_names'}")
    print(f"  {'-' * 30} {'-' * 15} {'-' * 15}  {'-' * 50}")
    for r in rows:
        ds_str = ", ".join(r["dataset_names"])
        print(f"  {r['robot_type']:<30s} {r['episode_number']:>15,} {r['total_hours']:>14.2f}h  {ds_str}")
    print(f"  {'-' * 30} {'-' * 15} {'-' * 15}")
    print(f"  {'TOTAL':<30s} {grand_episodes:>15,} {grand_hours:>14.2f}h")
    print(f"{'=' * 120}")
    if errors:
        print(f"  Errors: {len(errors)}")
        for e in errors[:10]:
            print(f"    {e['path']}: {e['error']}")
        if len(errors) > 10:
            print(f"    ... and {len(errors) - 10} more")
    print(f"  Elapsed: {elapsed:.1f}s\n")

    print("[3/3] Exporting xlsx ...")
    export_xlsx(rows, grand_episodes, grand_hours, grand_frames, args.output)
    print(f"  Output: {args.output}")
    print("Done.")


def export_xlsx(
    rows: list[dict],
    grand_episodes: int,
    grand_hours: float,
    grand_frames: int,
    output_path: str,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Robot Type Stats"

    headers = ["robot_type", "episode_number", "total_times(hours)", "dataset_names"]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, r in enumerate(rows, 2):
        ds_str = "\n".join(r["dataset_names"])
        values = [
            r["robot_type"],
            r["episode_number"],
            r["total_hours"],
            ds_str,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 4:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col_idx >= 2:
                cell.alignment = Alignment(horizontal="center")

    sum_row = len(rows) + 2
    ws.cell(row=sum_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=sum_row, column=2, value=grand_episodes).alignment = Alignment(horizontal="center")
    ws.cell(row=sum_row, column=3, value=grand_hours).alignment = Alignment(horizontal="center")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=sum_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.border = thin_border

    col_widths = [30, 20, 20, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    main()
