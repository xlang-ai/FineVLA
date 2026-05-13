#!/usr/bin/env python3
"""Calculate dataset durations for LeRobot v2.0 / v2.1 datasets.

Reads meta/info.json files from each dataset (or sub-dataset) and computes
duration as total_frames / fps. Sub-dataset durations are summed to produce
the total duration for each top-level dataset.

Usage:
    python calc_video_duration.py \
        --roots /path/to/Lerobot_v20 /path/to/Lerobot_v21 \
        --output dataset_durations.json
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from multiprocessing import cpu_count
from pathlib import Path



def format_duration(seconds: float) -> str:
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    if hours > 0:
        return f"{hours}h {minutes}m {secs}s"
    if minutes > 0:
        return f"{minutes}m {secs}s"
    return f"{secs}s"


SKIP_DIRS = {"videos", "data", "images", "logs", ".git", "__pycache__"}


def find_info_jsons(directory: str) -> list[str]:
    """Recursively find all meta/info.json files under a directory.

    Skips heavy directories (videos/, data/, etc.) that never contain
    meta/info.json to avoid traversing millions of video/parquet files.
    """
    results: list[str] = []
    for dirpath, dirnames, filenames in os.walk(directory):
        # Prune heavy dirs in-place so os.walk won't descend into them
        dirnames[:] = [
            d for d in dirnames
            if d not in SKIP_DIRS and not d.startswith("chunk-")
        ]
        if os.path.basename(dirpath) == "meta" and "info.json" in filenames:
            results.append(os.path.join(dirpath, "info.json"))
    return results


def read_info_json(filepath: str) -> dict:
    """Read an info.json and extract total_frames and fps.

    Returns dict with keys: path, total_frames, fps, duration_seconds, error.
    """
    result = {
        "path": filepath,
        "total_frames": 0,
        "fps": 0,
        "duration_seconds": 0.0,
        "error": "",
    }
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            info = json.load(f)

        total_frames = info.get("total_frames")
        fps = info.get("fps")

        if total_frames is None or fps is None:
            result["error"] = f"missing total_frames or fps"
            return result

        total_frames = int(total_frames)
        fps = float(fps)

        if fps <= 0:
            result["error"] = f"invalid fps={fps}"
            return result

        result["total_frames"] = total_frames
        result["fps"] = fps
        result["duration_seconds"] = total_frames / fps

    except Exception as e:
        result["error"] = str(e)

    return result


def process_dataset(ds_name: str, ds_path: str, version: str) -> dict:
    """Process a single top-level dataset: find all info.json, compute total duration.

    Files are read one at a time to avoid hitting the open-file-descriptor limit.
    """
    info_files = find_info_jsons(ds_path)

    total_duration = 0.0
    total_frames_all = 0
    errors = []

    for info_path in info_files:
        for attempt in range(3):
            r = read_info_json(info_path)
            if r["error"] and "Too many open files" in r["error"]:
                time.sleep(0.5)
                continue
            break
        if r["error"]:
            errors.append({"path": r["path"], "error": r["error"]})
        else:
            total_duration += r["duration_seconds"]
            total_frames_all += r["total_frames"]

    return {
        "name": ds_name,
        "path": ds_path,
        "version": version,
        "sub_dataset_count": len(info_files),
        "total_frames": total_frames_all,
        "total_duration_seconds": round(total_duration, 2),
        "total_duration_hours": round(total_duration / 3600, 4),
        "total_duration_readable": format_duration(total_duration),
        "error_count": len(errors),
        "errors": errors if errors else [],
    }


IGNORE_DATASETS = {"IPEC-COMMUNITY","RoboCOIN_annotations_backup"}


def scan_datasets(roots: list[str]) -> list[tuple[str, str, str]]:
    """Discover all direct sub-directories under each root.

    Returns list of (ds_name, ds_path, version).
    """
    datasets = []
    for root in roots:
        root_path = Path(root)
        if not root_path.exists():
            print(f"  Warning: {root} does not exist, skipping")
            continue
        version = root_path.name
        for ds_dir in sorted(root_path.iterdir()):
            if not ds_dir.is_dir():
                continue
            if ds_dir.name in IGNORE_DATASETS:
                print(f"  Skipping ignored dataset: {version}/{ds_dir.name}")
                continue
            ds_name = f"{version}/{ds_dir.name}"
            datasets.append((ds_name, str(ds_dir), version))
    return datasets


def main():
    parser = argparse.ArgumentParser(
        description="Calculate dataset durations from info.json (total_frames/fps)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--roots", nargs="+", required=True,
        help="Root directories (e.g. Lerobot_v20 Lerobot_v21)",
    )
    parser.add_argument(
        "--num-workers", type=int, default=None,
        help="Thread count for parallel processing. Default: cpu_count.",
    )
    parser.add_argument(
        "--output", type=str, required=True,
        help="Output JSON file path",
    )
    parser.add_argument(
        "--xlsx", action="store_true",
        help="Export datasets to .xlsx table and generate a pie chart image",
    )
    args = parser.parse_args()

    num_workers = args.num_workers or min(cpu_count(), 4)
    wall_start = time.time()

    # 1. Discover datasets
    print(f"[1/2] Scanning dataset directories ...")
    datasets = scan_datasets(args.roots)
    if not datasets:
        print("No datasets found. Exiting.")
        sys.exit(1)
    print(f"  Found {len(datasets)} top-level datasets\n")

    # 2. Process each dataset in parallel
    print(f"[2/2] Reading info.json files ({num_workers} workers) ...")

    ds_entries: list[dict] = []
    with ThreadPoolExecutor(max_workers=min(len(datasets), num_workers)) as pool:
        futures = {
            pool.submit(process_dataset, name, path, ver): name
            for name, path, ver in datasets
        }
        for fut in as_completed(futures):
            entry = fut.result()
            ds_entries.append(entry)
            status = ""
            if entry["error_count"] > 0:
                status = f"  ({entry['error_count']} errors)"
            print(
                f"  {entry['name']:<50s} "
                f"{entry['sub_dataset_count']:>5} info.json  "
                f"{entry['total_duration_readable']:>12s}{status}"
            )

    ds_entries.sort(key=lambda x: x["name"])

    elapsed = time.time() - wall_start

    # 3. Build output
    grand_total_seconds = sum(e["total_duration_seconds"] for e in ds_entries)
    grand_total_frames = sum(e["total_frames"] for e in ds_entries)
    grand_total_errors = sum(e["error_count"] for e in ds_entries)
    grand_total_sub = sum(e["sub_dataset_count"] for e in ds_entries)

    for entry in ds_entries:
        if not entry["errors"]:
            del entry["errors"]

    output = {
        "summary": {
            "total_datasets": len(ds_entries),
            "total_sub_datasets": grand_total_sub,
            "total_frames": grand_total_frames,
            "total_duration_seconds": round(grand_total_seconds, 2),
            "total_duration_hours": round(grand_total_seconds / 3600, 4),
            "total_duration_readable": format_duration(grand_total_seconds),
            "total_errors": grand_total_errors,
            "scan_roots": args.roots,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "num_workers": num_workers,
            "elapsed_seconds": round(elapsed, 2),
            "elapsed_readable": format_duration(elapsed),
            "method": "total_frames / fps from meta/info.json",
        },
        "datasets": ds_entries,
    }

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    # 4. Print summary
    print(f"\n{'=' * 78}")
    print("  Dataset Duration Statistics (from info.json: total_frames / fps)")
    print(f"{'=' * 78}")
    print(f"  Total datasets     : {len(ds_entries)}")
    print(f"  Total sub-datasets : {grand_total_sub}")
    print(f"  Total frames       : {grand_total_frames:,}")
    print(
        f"  Total duration     : {format_duration(grand_total_seconds)} "
        f"({grand_total_seconds / 3600:.2f} hours)"
    )
    print(f"  Total errors       : {grand_total_errors}")
    print(f"  Elapsed time       : {format_duration(elapsed)}")
    print(f"  Output file        : {args.output}")
    print(f"{'=' * 78}")
    for ds in ds_entries:
        line = (
            f"  {ds['name']:<50s} "
            f"{ds['sub_dataset_count']:>5} subs  "
            f"{ds['total_duration_readable']:>14s}"
        )
        if ds["error_count"]:
            line += f"  ({ds['error_count']} errors)"
        print(line)
    print(f"{'=' * 78}\n")

    # 5. Optional: export xlsx + pie chart
    if args.xlsx:
        export_xlsx_and_chart(ds_entries, grand_total_seconds, out_path)


def export_xlsx_and_chart(
    ds_entries: list[dict],
    grand_total_seconds: float,
    json_out_path: Path,
) -> None:
    """Export dataset stats to .xlsx and generate a duration pie chart."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    stem = json_out_path.stem
    parent = json_out_path.parent
    xlsx_path = parent / f"{stem}.xlsx"
    chart_path = parent / f"{stem}_pie.png"

    # ── xlsx ──────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Dataset Durations"

    headers = [
        "Dataset Name",
        "Version",
        "Sub-datasets",
        "Total Frames",
        "Duration (seconds)",
        "Duration (hours)",
        "Duration (readable)",
        "Errors",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, ds in enumerate(ds_entries, 2):
        values = [
            ds["name"],
            ds["version"],
            ds["sub_dataset_count"],
            ds["total_frames"],
            ds["total_duration_seconds"],
            ds["total_duration_hours"],
            ds["total_duration_readable"],
            ds["error_count"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal="center")

    # summary row
    sum_row = len(ds_entries) + 2
    ws.cell(row=sum_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=sum_row, column=3, value=sum(d["sub_dataset_count"] for d in ds_entries))
    ws.cell(row=sum_row, column=4, value=sum(d["total_frames"] for d in ds_entries))
    ws.cell(row=sum_row, column=5, value=round(grand_total_seconds, 2))
    ws.cell(row=sum_row, column=6, value=round(grand_total_seconds / 3600, 4))
    ws.cell(row=sum_row, column=7, value=format_duration(grand_total_seconds))
    ws.cell(row=sum_row, column=8, value=sum(d["error_count"] for d in ds_entries))
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=sum_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.border = thin_border
        if col_idx >= 3:
            cell.alignment = Alignment(horizontal="center")

    col_widths = [40, 14, 14, 16, 18, 16, 18, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(str(xlsx_path))
    print(f"  xlsx saved: {xlsx_path}")

    # ── pie chart（饼图中将 RoboCOIN 开头的多个数据集合并为一项 "RoboCOIN"）────────────────
    robocoin_entries = [
        e for e in ds_entries
        if e["name"].split("/", 1)[-1].startswith("RoboCOIN")
    ]
    other_entries = [e for e in ds_entries if e not in robocoin_entries]

    chart_names: list[str] = []
    chart_hours: list[float] = []
    if robocoin_entries:
        chart_names.append("Lerobot_v21/RoboCOIN")
        chart_hours.append(sum(e["total_duration_hours"] for e in robocoin_entries))
    for e in other_entries:
        chart_names.append(e["name"])
        chart_hours.append(e["total_duration_hours"])

    sorted_pairs = sorted(zip(chart_hours, chart_names), reverse=True)
    hours_sorted = [h for h, _ in sorted_pairs]
    names_sorted = [n for _, n in sorted_pairs]

    cmap = plt.get_cmap("tab20")
    colors = [cmap(i / len(names_sorted)) for i in range(len(names_sorted))]

    fig, ax = plt.subplots(figsize=(14, 8))
    fig.subplots_adjust(left=0.05, right=0.55)

    wedges, texts, autotexts = ax.pie(
        hours_sorted,
        labels=None,
        autopct=lambda pct: f"{pct:.1f}%" if pct >= 3 else "",
        startangle=140,
        colors=colors,
        pctdistance=0.8,
        wedgeprops={"edgecolor": "white", "linewidth": 1.2},
    )

    for t in autotexts:
        t.set_fontsize(8)
        t.set_fontweight("bold")

    legend_labels = [
        f"{n}  ({h:.1f}h)" for n, h in zip(names_sorted, hours_sorted)
    ]
    ax.legend(
        wedges,
        legend_labels,
        title="Datasets",
        title_fontsize=11,
        fontsize=9,
        loc="center left",
        bbox_to_anchor=(1.02, 0.5),
        frameon=True,
        fancybox=True,
        shadow=True,
    )

    total_hours = sum(chart_hours)
    ax.set_title(
        f"Dataset Duration Distribution (Total: {total_hours:.1f} hours)",
        fontsize=14,
        fontweight="bold",
        pad=20,
    )

    fig.savefig(str(chart_path), dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  pie chart saved: {chart_path}")


if __name__ == "__main__":
    main()
